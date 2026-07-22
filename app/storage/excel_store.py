"""Excel (``.xlsx``) implementation of :class:`~app.storage.base.PropertyStore`.

Design goals (see the redesign brief):

* **Dynamic schema** - columns are matched by header *name*, never by index.
  Writing a field that has no column yet appends the column automatically, so
  new fields never break existing files.
* **Concurrency-safe** - every read-modify-write is wrapped in a
  :class:`~app.storage.locking.FileLock`, so multiple threads/processes can
  append and edit without corrupting the workbook.
* **Self-healing** - the workbook (and worksheet) is created on first write if
  it does not exist yet.

Only openpyxl is required; there is no Google/network dependency.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core.exceptions import SheetError
from app.core.logging import get_logger
from app.storage.base import PropertyStore, StoredRecord
from app.storage.locking import FileLock

logger = get_logger(__name__)

#: Row 1 holds the header; data starts at row 2.
_HEADER_ROW = 1
_FIRST_DATA_ROW = 2


def _clean(value: Any) -> Any:
    """Normalise a cell value: ``None`` -> ``""`` and trim header strings."""
    return "" if value is None else value


class ExcelPropertyStore(PropertyStore):
    """A property database stored in a single ``.xlsx`` worksheet."""

    def __init__(
        self,
        path: str | Path,
        worksheet: str = "Properties",
        *,
        lock_timeout: float = 30.0,
    ) -> None:
        self._path = Path(path)
        self._worksheet = worksheet
        self._lock = FileLock(self._path, timeout=lock_timeout)

    @property
    def path(self) -> Path:
        return self._path

    # ------------------------------------------------------------- workbook IO

    def _ensure_file(self) -> None:
        """Create an empty workbook with the target worksheet if missing."""
        if self._path.exists():
            return
        self._path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = self._worksheet
        wb.save(self._path)
        wb.close()
        logger.info("Created new property workbook at %s", self._path)

    def _open(self, *, read_only: bool = False) -> tuple[Any, Worksheet]:
        try:
            wb = load_workbook(self._path, data_only=read_only)
        except Exception as exc:  # openpyxl raises many concrete types
            raise SheetError(f"Failed to open workbook {self._path}: {exc}") from exc
        ws = wb[self._worksheet] if self._worksheet in wb.sheetnames else wb.active
        return wb, ws

    @staticmethod
    def _read_header(ws: Worksheet) -> list[str]:
        if ws.max_row < _HEADER_ROW:
            return []
        return [str(_clean(c.value)).strip() for c in ws[_HEADER_ROW]]

    @staticmethod
    def _row_to_values(header: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
        return {
            header[i]: _clean(row[i]) if i < len(row) else ""
            for i in range(len(header))
            if header[i]
        }

    # --------------------------------------------------------------- reads

    def headers(self) -> list[str]:
        if not self._path.exists():
            return []
        with self._lock:
            wb, ws = self._open(read_only=True)
            try:
                return [h for h in self._read_header(ws) if h]
            finally:
                wb.close()

    def read_all(self) -> list[StoredRecord]:
        if not self._path.exists():
            return []
        with self._lock:
            wb, ws = self._open(read_only=True)
            try:
                header = self._read_header(ws)
                if not any(header):
                    return []
                records: list[StoredRecord] = []
                for row_id, row in enumerate(
                    ws.iter_rows(min_row=_FIRST_DATA_ROW, values_only=True),
                    start=_FIRST_DATA_ROW,
                ):
                    if row is None or all(_clean(v) == "" for v in row):
                        continue
                    records.append(StoredRecord(row_id=row_id, values=self._row_to_values(header, row)))
                return records
            finally:
                wb.close()

    def get(self, row_id: int) -> StoredRecord | None:
        for record in self.read_all():
            if record.row_id == row_id:
                return record
        return None

    def find_by(self, column: str, value: Any) -> StoredRecord | None:
        target = str(_clean(value)).strip().lower()
        for record in self.read_all():
            if str(record.get(column, "")).strip().lower() == target:
                return record
        return None

    def search(self, query: str, columns: list[str] | None = None) -> list[StoredRecord]:
        needle = (query or "").strip().lower()
        if not needle:
            return self.read_all()
        matches: list[StoredRecord] = []
        for record in self.read_all():
            fields = columns or list(record.values.keys())
            haystack = " ".join(str(record.get(c, "")) for c in fields).lower()
            if needle in haystack:
                matches.append(record)
        return matches

    # --------------------------------------------------------------- writes

    def ensure_columns(self, names: list[str]) -> None:
        wanted = [n for n in names if n]
        if not wanted:
            return
        with self._lock:
            self._ensure_file()
            wb, ws = self._open()
            try:
                if self._add_missing_columns(ws, wanted):
                    wb.save(self._path)
            finally:
                wb.close()

    @staticmethod
    def _add_missing_columns(ws: Worksheet, names: list[str]) -> bool:
        """Append any missing headers to row 1. Returns True if changed."""
        header = [str(_clean(c.value)).strip() for c in ws[_HEADER_ROW]] if ws.max_row >= 1 else []
        header = [h for h in header if h]
        existing = {h.lower() for h in header}
        added = False
        for name in names:
            if name.lower() not in existing:
                header.append(name)
                existing.add(name.lower())
                ws.cell(row=_HEADER_ROW, column=len(header), value=name)
                added = True
        return added

    def append(self, values: dict[str, Any]) -> StoredRecord:
        with self._lock:
            self._ensure_file()
            wb, ws = self._open()
            try:
                self._add_missing_columns(ws, list(values.keys()))
                header = [str(_clean(c.value)).strip() for c in ws[_HEADER_ROW]]
                index = {h.lower(): i + 1 for i, h in enumerate(header) if h}
                # First empty row after the last used row.
                new_row = max(ws.max_row, _HEADER_ROW) + 1
                for key, value in values.items():
                    col = index.get(key.lower())
                    if col is not None:
                        ws.cell(row=new_row, column=col, value=_serialise(value))
                wb.save(self._path)
                stored = {h: values.get(h, "") for h in header if h}
                logger.info("Appended property to %s at row %d", self._path.name, new_row)
                return StoredRecord(row_id=new_row, values=stored)
            finally:
                wb.close()

    def update(self, row_id: int, values: dict[str, Any]) -> StoredRecord:
        with self._lock:
            self._ensure_file()
            wb, ws = self._open()
            try:
                if row_id < _FIRST_DATA_ROW or row_id > ws.max_row:
                    raise SheetError(f"Row {row_id} is out of range in {self._path.name}")
                self._add_missing_columns(ws, list(values.keys()))
                header = [str(_clean(c.value)).strip() for c in ws[_HEADER_ROW]]
                index = {h.lower(): i + 1 for i, h in enumerate(header) if h}
                for key, value in values.items():
                    col = index.get(key.lower())
                    if col is not None:
                        ws.cell(row=row_id, column=col, value=_serialise(value))
                wb.save(self._path)
                current = self._row_to_values(
                    header, tuple(c.value for c in ws[row_id])
                )
                return StoredRecord(row_id=row_id, values=current)
            finally:
                wb.close()


def _serialise(value: Any) -> Any:
    """Coerce values Excel cannot store natively (lists) to strings."""
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(v) for v in value)
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return value
