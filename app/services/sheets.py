"""Spreadsheet data source.

Supports two interchangeable backends behind one interface:

- ``GoogleSheetSource`` - Google Sheets via a service account (gspread)
- ``ExcelSheetSource``  - a local .xlsx file (openpyxl)

Both read rows into ``PropertyData`` and can write back Status, Title,
Description, Listing URL, Error and Updated Date for a given row.
"""

from __future__ import annotations

from abc import ABC, abstractmethod
from datetime import UTC, datetime
from pathlib import Path

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

from app.core.config import APP_DIR, SheetConfig, get_config
from app.core.exceptions import SheetError
from app.core.logging import get_logger
from app.models.schemas import SHEET_COLUMNS, PropertyData

logger = get_logger(__name__)

_SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# Columns the app is allowed to write back to the sheet
WRITABLE_COLUMNS = ["Status", "Title", "Description", "Listing URL", "Error", "Updated Date"]


class SheetSource(ABC):
    """Abstract spreadsheet backend."""

    @abstractmethod
    def read_properties(self) -> list[PropertyData]:
        """Read all rows as validated PropertyData objects."""

    @abstractmethod
    def write_back(self, sheet_row: int, values: dict[str, str]) -> None:
        """Write selected columns back to one row (1-based sheet row number)."""

    def mark_posted(self, sheet_row: int, listing_url: str, title: str, description: str) -> None:
        """Convenience: mark a row Posted with its listing URL."""
        self.write_back(
            sheet_row,
            {
                "Status": "Posted",
                "Title": title,
                "Description": description,
                "Listing URL": listing_url,
                "Error": "",
                "Updated Date": datetime.now(UTC).strftime("%Y-%m-%d %H:%M"),
            },
        )

    def mark_failed(self, sheet_row: int, error: str) -> None:
        """Convenience: record a failure on the row."""
        self.write_back(
            sheet_row,
            {
                "Status": "Failed",
                "Error": error[:500],
                "Updated Date": datetime.now(UTC).strftime("%Y-%m-%d %H:%M"),
            },
        )


class GoogleSheetSource(SheetSource):
    """Google Sheets backend using a service account."""

    def __init__(self, config: SheetConfig | None = None) -> None:
        self._config = config or get_config().sheet
        if not self._config.spreadsheet_id:
            raise SheetError("sheet.spreadsheet_id is not configured")
        sa_path = Path(self._config.service_account_file)
        if not sa_path.is_absolute():
            sa_path = APP_DIR / sa_path
        if not sa_path.exists():
            raise SheetError(f"Service account file not found: {sa_path}")
        try:
            creds = Credentials.from_service_account_file(str(sa_path), scopes=_SCOPES)
            client = gspread.authorize(creds)
            self._sheet = client.open_by_key(self._config.spreadsheet_id).worksheet(
                self._config.worksheet_name
            )
        except Exception as exc:  # gspread raises many concrete types
            raise SheetError(f"Failed to open Google Sheet: {exc}") from exc
        self._header: list[str] = []

    def _load_header(self) -> list[str]:
        if not self._header:
            self._header = [h.strip() for h in self._sheet.row_values(1)]
        return self._header

    def read_properties(self) -> list[PropertyData]:
        try:
            records = self._sheet.get_all_records()
        except Exception as exc:
            raise SheetError(f"Failed to read Google Sheet: {exc}") from exc
        properties: list[PropertyData] = []
        for i, record in enumerate(records, start=2):  # row 1 is the header
            if not str(record.get("Property ID", "")).strip():
                continue
            try:
                properties.append(PropertyData.from_sheet_row(record, sheet_row=i))
            except Exception as exc:  # noqa: BLE001 - one bad row must not stop the sync
                logger.error("Skipping invalid sheet row %d: %s", i, exc)
        logger.info("Read %d properties from Google Sheet", len(properties))
        return properties

    def write_back(self, sheet_row: int, values: dict[str, str]) -> None:
        header = self._load_header()
        updates = []
        for column, value in values.items():
            if column not in WRITABLE_COLUMNS or column not in header:
                continue
            col_index = header.index(column) + 1
            updates.append(gspread.cell.Cell(row=sheet_row, col=col_index, value=value))
        if not updates:
            return
        try:
            self._sheet.update_cells(updates)
        except Exception as exc:
            raise SheetError(f"Failed to write back to Google Sheet row {sheet_row}: {exc}") from exc


class ExcelSheetSource(SheetSource):
    """Local Excel (.xlsx) backend."""

    def __init__(self, config: SheetConfig | None = None) -> None:
        self._config = config or get_config().sheet
        self._path = Path(self._config.excel_path)
        if not self._path.exists():
            raise SheetError(f"Excel file not found: {self._path}")

    def read_properties(self) -> list[PropertyData]:
        try:
            wb = load_workbook(self._path, data_only=True)
            ws = wb[self._config.worksheet_name] if self._config.worksheet_name in wb.sheetnames else wb.active
        except Exception as exc:
            raise SheetError(f"Failed to open Excel file: {exc}") from exc

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(h or "").strip() for h in rows[0]]
        properties: list[PropertyData] = []
        for i, row in enumerate(rows[1:], start=2):
            record = {header[j]: (row[j] if j < len(row) else "") for j in range(len(header))}
            record = {k: ("" if v is None else v) for k, v in record.items()}
            if not str(record.get("Property ID", "")).strip():
                continue
            try:
                properties.append(PropertyData.from_sheet_row(record, sheet_row=i))
            except Exception as exc:  # noqa: BLE001
                logger.error("Skipping invalid Excel row %d: %s", i, exc)
        wb.close()
        logger.info("Read %d properties from Excel", len(properties))
        return properties

    def write_back(self, sheet_row: int, values: dict[str, str]) -> None:
        try:
            wb = load_workbook(self._path)
            ws = wb[self._config.worksheet_name] if self._config.worksheet_name in wb.sheetnames else wb.active
            header = [str(c.value or "").strip() for c in ws[1]]
            for column, value in values.items():
                if column in WRITABLE_COLUMNS and column in header:
                    ws.cell(row=sheet_row, column=header.index(column) + 1, value=value)
            wb.save(self._path)
            wb.close()
        except Exception as exc:
            raise SheetError(f"Failed to write back to Excel row {sheet_row}: {exc}") from exc


def create_sheet_source(config: SheetConfig | None = None) -> SheetSource:
    """Factory choosing the backend from configuration."""
    cfg = config or get_config().sheet
    if cfg.source_type == "google":
        return GoogleSheetSource(cfg)
    return ExcelSheetSource(cfg)


# Ordered header row for generating example sheets
EXAMPLE_HEADER: list[str] = list(SHEET_COLUMNS.keys())
